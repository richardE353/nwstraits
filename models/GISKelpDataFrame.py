import os
from dataclasses import dataclass
from pathlib import Path

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pandas import DataFrame

import utils.FilesHelper as fh

from models.KelpDataRow import KelpSurvey


@dataclass
class GISData:
    survey_date: str
    surveyor: str
    county: str
    miles: float
    start_time: str
    end_time: str
    weather: str
    observations: str
    depth1_at_shore_edge: float
    depth1_at_outer_edge: float
    temp1_at_shore_edge: float
    temp1_at_outer_edge: float
    depth2_at_shore_edge: float
    depth2_at_outer_edge: float
    temp2_at_shore_edge: float
    temp2_at_outer_edge: float
    location: str
    survey_notes: str
    tidal_height: float
    tide_station: str
    survey_id_string: str
    mllw_tidal_ht: float
    mllw_depth1_at_shore_edge: float
    mllw_depth1_at_outer_edge: float
    mllw_depth2_at_shore_edge: float
    mllw_depth2_at_outer_edge: float
    to_beach_url: str
    survey_conditions: str
    to_beach_file: str
    current_knots: float
    current_station: str
    extent_start_waypoint: str
    extent_end_waypoint: str


def as_gis_data(ks: KelpSurvey) -> GISData:
    da = ks.depth_adjuster
    return GISData(
        ks.survey_date,
        ks.volunteer_info.lead_name,
        ks.county,
        0.0,  # niles
        ks.time_start[:5],
        ks.time_end[:5],
        ks.weather,
        ks.survey_observations,
        round(ks.depth_1_at_shore_edge, 2),
        round(ks.depth_1_at_outer_edge, 2),
        round(ks.temp_1_at_shore_edge, 1),
        round(ks.temp_1_at_outer_edge, 1),
        round(ks.depth_2_at_shore_edge, 2),
        round(ks.depth_2_at_outer_edge, 2),
        round(ks.temp_2_at_shore_edge, 1),
        round(ks.temp_2_at_outer_edge, 1),
        ks.location,
        ks.survey_notes,
        round(ks.tidal_ht, 2),
        ks.tide_station_label,
        ks.file_prefix(),
        round(da.adjust_depth(ks.tidal_ht), 2),
        round(da.adjust_depth(ks.depth_1_at_shore_edge), 2),
        round(da.adjust_depth(ks.depth_1_at_outer_edge), 2),
        round(da.adjust_depth(ks.depth_2_at_shore_edge), 2),
        round(da.adjust_depth(ks.depth_2_at_outer_edge), 2),
        'no image available',
        ks.survey_conditions,
        file_hyperlink(ks.site_image_names.to_beach),
        round(ks.current_knots, 1),
        ks.current_station,
        ks.extent_start_waypoint,
        ks.extent_end_waypoint
    )


def row_header_labels() -> list:
    return ["Survey_Date",
            "Surveyor",
            "County",
            "Miles",
            "Start_Time",
            "End_Time",
            "Weather",
            "Observations",
            "D1shore_Edge",
            "D1water_Edge",
            "T1ShoreEdge",
            "T1WaterEdge",
            "D2shore_Edge",
            "D2water_Edge",
            "T2ShoreEdge",
            "T2WaterEdge",
            "Bedname",
            "Additional_Obs",
            "Tidal_Ht",
            "Tide_Station",
            "Survey_Id_String",
            "MLLW_Tidal_Ht_meters",
            "MLLW_D1shore_meters",
            "MLLW_D1water_meters",
            "MLLW_D2shore_meters",
            "MLLW_D2water_meters",
            "ToBe",
            "survey_conditions",
            "ToBe_file",
            "current_knots",
            "current_station",
            "extent_start_waypoint",
            "extent_end_waypoint"
            ]


def file_hyperlink(fn: str) -> str:
    if len(fn) == 0:
        return ''
    file_path = fh.find_attachment_path(fn)
    label = Path(file_path).name

    return path_to_link(file_path, label)


# https://trumpexcel.com/hyperlinks/
def path_to_link(path, label):
    return '=HYPERLINK("{}","{}")'.format(path, label)


# https://www.kdnuggets.com/2022/08/3-ways-append-rows-pandas-dataframes.html
# https://groups.google.com/g/openpyxl-users/c/1auXBiDlzHk?pli=1 (final comment)
def create_gis_excel_workbook(surveys: dict, dest: str):
    gis_data = []
    for cty in surveys.keys():
        for ks in surveys[cty]:
            gis_data.append(as_gis_data(ks))

    df = DataFrame(gis_data)

    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    set_column_labels(ws)

    number_cols = ['D', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'S', 'V', 'W', 'X', 'Y', 'Z', 'AD', 'AF', 'AG']
    set_number_formats(ws, number_cols, '0.00')

    destination = str(os.path.join(dest, 'gisWorksheet.xlsx'))
    print('\t\twriting ', destination)

    wb.save(destination)


# See 'Converting worksheet to a Dataframe' on https://openpyxl.readthedocs.io/en/stable/pandas.html
def as_data_frame(sheet) -> DataFrame:
    from itertools import islice

    data = sheet.values
    cols = next(data)[1:]
    data = list(data)
    idx = [r[0] for r in data]
    data = (islice(r, 1, None) for r in data)
    return DataFrame(data, index=idx, columns=cols)


def set_number_formats(sheet: Worksheet, cell_letters: list, fmt: str):
    for c in cell_letters:
        for i in range(1, sheet.max_row):
            sheet[c + str(i)].number_format = fmt

def set_column_labels(sheet: Worksheet):
    labels = row_header_labels()
    for i in range(1, sheet.max_column):
        sheet.cell(1, i).value = labels[i - 1]
