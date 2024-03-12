import os
import sys

import openpyxl
from openpyxl.workbook import Workbook

import RuntimeArgs as rt_args

from pandas import DataFrame

from SmugMug import update_urls, images_for_year, resize_thumbnail
from models.GISKelpDataFrame import as_data_frame

years = [2019, 2020, 2021]


# uncomment the script you want to run
def main():
    # dump_smugmug_urls()
    update_attribute_image_urls()


def dump_smugmug_urls():
    dfs = create_image_dataframes(years)
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    y_index = 0
    for year in years:
        ws = wb.create_sheet(title=str(year), index=y_index)
        df = dfs[y_index]
        df.sort_values(by=['fileName'], inplace=True)
        y_index = y_index + 1

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    ws_path = os.path.join(rt_args.default_output_dir, 'smugmugLinks.xlsx')
    wb.save(str(ws_path))


def update_attribute_image_urls():
    dfs = create_image_dataframes(years)

    new_image_urls = {}
    for df in dfs:
        for i_url in df['xlURL']:
            image_id = extract_image_id(i_url, 5)
            new_image_urls[image_id] = i_url

    old_data_path = os.path.join(rt_args.default_output_dir, 'SIQ_Kelp_2019thru2021_attribute_table_old_urls.xlsx')
    wb = openpyxl.load_workbook(old_data_path, rich_text=True)
    valid_sheets = list(filter(lambda s: 'SIQ_Kelp' in s, wb.sheetnames))

    for sn in valid_sheets:
        sheet = wb[sn]
        old_df = as_data_frame(sheet)

        for c in ['ToBe', 'ToWa', 'BeL', 'BeR']:
            cur_col = old_df[c]
            new_urls = []
            for old_url in cur_col:
                if old_url:
                    image_id = extract_image_id(old_url, 2)
                    if image_id in new_image_urls:
                        n_url = new_image_urls[image_id]
                        new_urls.append(n_url)
                    else:
                        new_urls.append(old_url)
                        print("failed to find match for {}".format(old_url))
                else:
                    new_urls.append(None)

            update_urls(sheet, c, new_urls)

    out_dir = os.path.join(rt_args.default_output_dir, 'new_urls.xlsx')
    wb.save(str(out_dir))


def extract_image_id(url: str, index_from_r: int = 2) -> str:
    chunks = url.rsplit('/')
    if chunks:
        return chunks[len(chunks) - index_from_r]

    return ''


def create_image_dataframes(selected_years: list) -> list:
    dataframes = map(create_image_dataframe, selected_years)
    return list(dataframes)


def create_image_dataframe(year: int) -> DataFrame:
    all_images = images_for_year(year)
    image_rows = []

    for image in all_images:
        xl_url = resize_thumbnail(image['ThumbnailUrl'], 'XL')

        new_row = {'fileName': image['FileName'],
                   'thumbnailURL': image['ThumbnailUrl'],
                   'xlURL': xl_url}

        image_rows.append(new_row)

    return DataFrame(image_rows)


if __name__ == "__main__":
    sys.exit(main())
