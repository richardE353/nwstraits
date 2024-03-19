"""
Update 'ToBe' column in gisWorksheet to match associated SmugMug picture
"""
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import RuntimeArgs as rt_args
import sys

import pandas as pd
from pandas import DataFrame

import requests

from models.GISKelpDataFrame import as_data_frame


def main():
    try:
        year = rt_args.select_collection_year()
        src_dir = select_gis_worksheet_dir(year)

        print("\nRuntime parameters")
        print('\tYear: ' + str(year))
        print('\tGISWorksheet directory: ' + src_dir)

        print('\n\tgetting beach images from SmugMug')
        beach_images = beach_images_for_year(year)

        if beach_images:
            print('\tloading gisWorksheet')
            fn = 'gisWorksheet' + str(year) + '.xlsx'
            ws_path = os.path.join(src_dir, fn)
            wb = openpyxl.load_workbook(ws_path, rich_text=True)
            sheet = wb.active

            print('\tupdating {} ToBe urls'.format(len(beach_images)))

            urls = get_updated_to_beach_urls(beach_images, as_data_frame(sheet))
            update_urls(sheet, 'ToBe', urls)

            print('\twriting out GisWorksheet.xlsx')
            wb.save(ws_path)

        print('Done.')
    except Exception as e:
        print("Failed to update using SmugMug info due to: {}".format(e))


def get_updated_to_beach_urls(beach_images: list, gis_df: DataFrame) -> list:
    gis_records = gis_df.to_dict(orient='records')

    for i in beach_images:
        s_id = extract_survey_id_string(i['FileName'])
        rec = find_rec_with_survey_id(gis_records, s_id)

        if rec is not None:
            rec['ToBe'] = resize_thumbnail(i['ThumbnailUrl'], 'XL')

    return list(map(lambda r: r['ToBe'], gis_records))


def update_urls(sheet: Worksheet, col_lbl: str, urls: list):
    col_names = {}
    current = 1
    for c in sheet.iter_cols(1, sheet.max_column):
        col_names[c[0].value] = current
        current += 1

    cl = get_column_letter(col_names[col_lbl])
    cells = sheet[cl]
    for i in range(len(urls)):
        cells[i + 1].value = urls[i]


def images_for_year(year: int) -> list:
    y_node = get_node_for_year(year)

    if y_node:
        album_uri = y_node['Uris']['Album']['Uri']

        # https://api.smugmug.com/api/v2/doc/reference/image.html - see section on getting images for sharing
        url = 'https://api.smugmug.com' + album_uri + '!images?start=1&count=200'
        image_dicts = get_reply_from_url(url)['Response']['AlbumImage']
        return image_dicts

    print('\t***SmugMug node not found for year: {}'.format(year))
    return []


def beach_images_for_year(year: int) -> list:
    def is_beach(image: dict) -> bool:
        return '_ToBe.' in image['FileName']

    images = images_for_year(year)
    return list(filter(is_beach, images))


def get_node_for_year(year: int) -> dict:
    # The only way I've found to get the correct node is to "walk" the API, one node at a time.
    # since node values do not change, I did that using the SmugMug api web page, and have cached
    # the sound_iq_kelp_node.  The path I actually walked is as follows:
    # top_level_node = 'PS6Mc'  # top level node from https://api.smugmug.com/api/v2/user/nwstraits
    # projects_node = 'vNzZp'  # first node in https://api.smugmug.com/api/v2/node/PS6Mc!children
    # kelp_node = 'rTwcc'  # from https://api.smugmug.com/api/v2/node/vNzZp!children
    sound_iq_kelp_node = 'xVcZmc'  # from https://api.smugmug.com/api/v2/node/rTwcc!children

    node_url = 'https://api.smugmug.com/api/v2/node/' + sound_iq_kelp_node + '!children'
    rd = get_reply_from_url(node_url)

    yr_str = str(year)
    for n in rd['Response']['Node']:
        if n['Name'] == yr_str:
            return n


def get_reply_from_url(url: str) -> dict:
    sm_payload = {'APIKey': rt_args.smug_mug_key}
    sm_headers = {'Accept': 'application/json'}

    r = requests.get(url, params=sm_payload, headers=sm_headers)

    if r.status_code != 200:
        print("Query failed due to: {}", r.reason)
        print(r.content)

    return dict(r.json())


def import_gis_dataframe(src_path: str) -> DataFrame:
    # just using read_excel trashes formulas.
    # https://stackoverflow.com/questions/58160030/how-to-read-xlsx-as-pandas-dataframe-with-formulas-as-strings
    from openpyxl import load_workbook
    wb = load_workbook(filename=src_path)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)

    column_names = df.iloc[0]
    df = df[1:]
    df.columns = column_names

    return df


def extract_survey_id_string(fn: str) -> str:
    loc = fn.index('_ToBe')
    return fn[:loc]


def resize_thumbnail(url: str, to_sz: str) -> str:
    # 'ThumbnailUrl' contains URL for thumbnail size.
    #       image template is: https://photos.smugmug.com/photos/#image_idd#/0/#size#/i-6Vw7HKp-#size#.jpg
    #           where #size# in {'Ti', 'S', 'M', 'L', 'XL', and 'Th'}
    #       To see image sizes, use: https://api.smugmug.com/api/v2/image/#image_id#!sizedetails
    #           ex: https://api.smugmug.com/api/v2/image/6Vw7HKp-0!sizedetails
    new_url = url.replace('0/Th/', '0/' + to_sz + '/')
    return new_url.replace('-Th.', '-' + to_sz + '.')


def find_rec_with_survey_id(records: list, pattern: str) -> dict:
    for r in records:
        if r['Survey_Id_String'] == pattern:
            return r

    print('Failed to find survey_id_string: ', pattern)


def select_gis_worksheet_dir(year: int) -> str:
    output_dir = input("GISWorksheet directory: ").strip()
    if len(output_dir) > 0:
        return output_dir

    return os.path.join(rt_args.default_output_dir, str(year))


if __name__ == "__main__":
    sys.exit(main())
